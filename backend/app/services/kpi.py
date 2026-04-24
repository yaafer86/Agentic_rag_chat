"""KPI engine: safe formula evaluator + Excel table extraction.

The formula parser uses Python's AST so we can statically reject anything that
isn't a pure arithmetic expression over a small whitelist of names and numeric
literals. No `eval`, no attribute access, no calls to arbitrary functions.

Allowed:
  - binary ops: + - * / // % **
  - unary ops: + -
  - comparisons: < <= > >= == !=
  - boolean: and or not
  - constant numbers and strings
  - names bound to numeric values in the variable map
  - a fixed set of pure functions: min, max, sum, abs, round, avg
"""
from __future__ import annotations

import ast
import io
import logging
import math
import operator as op
from dataclasses import dataclass
from typing import Any

logger = logging.getLogger(__name__)


class FormulaError(ValueError):
    pass


_BIN_OPS: dict[type[ast.operator], Any] = {
    ast.Add: op.add,
    ast.Sub: op.sub,
    ast.Mult: op.mul,
    ast.Div: op.truediv,
    ast.FloorDiv: op.floordiv,
    ast.Mod: op.mod,
    ast.Pow: op.pow,
}

_UNARY_OPS: dict[type[ast.unaryop], Any] = {
    ast.UAdd: op.pos,
    ast.USub: op.neg,
    ast.Not: op.not_,
}

_CMP_OPS: dict[type[ast.cmpop], Any] = {
    ast.Lt: op.lt,
    ast.LtE: op.le,
    ast.Gt: op.gt,
    ast.GtE: op.ge,
    ast.Eq: op.eq,
    ast.NotEq: op.ne,
}


def _avg(xs):
    xs = list(xs)
    if not xs:
        return 0.0
    return sum(xs) / len(xs)


_FUNCS: dict[str, Any] = {
    "min": min,
    "max": max,
    "sum": lambda *args: sum(args) if args else 0,
    "abs": abs,
    "round": round,
    "avg": _avg,
    "sqrt": math.sqrt,
}


def evaluate_formula(formula: str, variables: dict[str, Any]) -> Any:
    """Safely evaluate `formula` against `variables`. Raises FormulaError on anything else."""
    try:
        tree = ast.parse(formula, mode="eval")
    except SyntaxError as e:
        raise FormulaError(f"invalid formula: {e}") from e
    return _eval(tree.body, variables)


def _eval(node: ast.AST, variables: dict[str, Any]) -> Any:
    if isinstance(node, ast.Constant):
        if isinstance(node.value, (int, float, str, bool)) or node.value is None:
            return node.value
        raise FormulaError(f"unsupported constant: {type(node.value).__name__}")
    if isinstance(node, ast.Name):
        if node.id in variables:
            return variables[node.id]
        raise FormulaError(f"unknown variable: {node.id}")
    if isinstance(node, ast.BinOp):
        f = _BIN_OPS.get(type(node.op))
        if not f:
            raise FormulaError(f"unsupported operator: {type(node.op).__name__}")
        return f(_eval(node.left, variables), _eval(node.right, variables))
    if isinstance(node, ast.UnaryOp):
        f = _UNARY_OPS.get(type(node.op))
        if not f:
            raise FormulaError(f"unsupported unary: {type(node.op).__name__}")
        return f(_eval(node.operand, variables))
    if isinstance(node, ast.BoolOp):
        values = [_eval(v, variables) for v in node.values]
        if isinstance(node.op, ast.And):
            return all(values)
        return any(values)
    if isinstance(node, ast.Compare):
        left = _eval(node.left, variables)
        for comp_op, right_node in zip(node.ops, node.comparators, strict=True):
            f = _CMP_OPS.get(type(comp_op))
            if not f:
                raise FormulaError(f"unsupported comparison: {type(comp_op).__name__}")
            right = _eval(right_node, variables)
            if not f(left, right):
                return False
            left = right
        return True
    if isinstance(node, ast.Call):
        if not isinstance(node.func, ast.Name):
            raise FormulaError("only named function calls are allowed")
        fn = _FUNCS.get(node.func.id)
        if not fn:
            raise FormulaError(f"unknown function: {node.func.id}")
        if node.keywords:
            raise FormulaError("keyword arguments not supported")
        args = [_eval(a, variables) for a in node.args]
        return fn(*args)
    if isinstance(node, ast.IfExp):
        return (
            _eval(node.body, variables)
            if _eval(node.test, variables)
            else _eval(node.orelse, variables)
        )
    raise FormulaError(f"unsupported expression: {type(node).__name__}")


# ------------------------------------------------------------------------------
# KPI validation (3 layers from the spec)
# ------------------------------------------------------------------------------


@dataclass
class ValidationIssue:
    layer: str
    level: str
    message: str


def validate_formula_shape(formula: str, variables: dict[str, Any]) -> list[ValidationIssue]:
    """Layer 1 — type / syntax check. Evaluates with sample values."""
    issues: list[ValidationIssue] = []
    try:
        value = evaluate_formula(formula, variables)
    except FormulaError as e:
        issues.append(ValidationIssue("syntax", "error", str(e)))
        return issues
    if not isinstance(value, (int, float, bool)):
        issues.append(
            ValidationIssue(
                "type",
                "warn",
                f"formula returned {type(value).__name__}; numeric/boolean expected",
            )
        )
    return issues


def validate_arithmetic_consistency(
    formula: str,
    cases: list[dict[str, Any]],
) -> list[ValidationIssue]:
    """Layer 2 — evaluate across several cases; flag NaN / overflow / division by zero."""
    issues: list[ValidationIssue] = []
    for i, case in enumerate(cases):
        try:
            v = evaluate_formula(formula, case)
        except ZeroDivisionError:
            issues.append(
                ValidationIssue("arithmetic", "error", f"case {i}: division by zero")
            )
            continue
        except FormulaError as e:
            issues.append(ValidationIssue("arithmetic", "error", f"case {i}: {e}"))
            continue
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            issues.append(ValidationIssue("arithmetic", "warn", f"case {i}: non-finite result {v}"))
    return issues


def validate_cross_source(
    expected_sources: list[str], actual_sources: list[str]
) -> list[ValidationIssue]:
    """Layer 3 — ensure the KPI references only known source documents."""
    actual = set(actual_sources)
    missing = [s for s in expected_sources if s not in actual]
    unknown = [s for s in actual_sources if s not in set(expected_sources)]
    issues: list[ValidationIssue] = []
    for m in missing:
        issues.append(ValidationIssue("cross_source", "error", f"missing source {m}"))
    for u in unknown:
        issues.append(ValidationIssue("cross_source", "warn", f"unknown source {u}"))
    return issues


# ------------------------------------------------------------------------------
# Excel table detection
# ------------------------------------------------------------------------------


@dataclass
class DetectedTable:
    sheet: str
    header: list[str]
    rows: list[list[Any]]
    row_offset: int
    col_offset: int


def detect_tables(xlsx_bytes: bytes) -> list[DetectedTable]:
    """Return all contiguous rectangular tables in each sheet of an xlsx file.

    Heuristic: the first row with >= 2 non-empty cells is taken as the header,
    then all subsequent rows that have at least one cell in the same column
    range are pulled in, stopping at the first fully-empty row.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    out: list[DetectedTable] = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        i = 0
        while i < len(rows):
            # Find header.
            while i < len(rows) and len([c for c in rows[i] if c not in (None, "")]) < 2:
                i += 1
            if i >= len(rows):
                break
            header_row = rows[i]
            header_cols = [j for j, c in enumerate(header_row) if c not in (None, "")]
            col_lo, col_hi = header_cols[0], header_cols[-1]
            header = [str(header_row[j]) for j in range(col_lo, col_hi + 1)]
            body: list[list[Any]] = []
            j = i + 1
            while j < len(rows):
                row = rows[j]
                slice_ = list(row[col_lo : col_hi + 1])
                if all(c in (None, "") for c in slice_):
                    break
                body.append(slice_)
                j += 1
            out.append(
                DetectedTable(
                    sheet=ws.title,
                    header=header,
                    rows=body,
                    row_offset=i,
                    col_offset=col_lo,
                )
            )
            i = j + 1
    return out


__all__ = [
    "DetectedTable",
    "FormulaError",
    "ValidationIssue",
    "detect_tables",
    "evaluate_formula",
    "validate_arithmetic_consistency",
    "validate_cross_source",
    "validate_formula_shape",
]
