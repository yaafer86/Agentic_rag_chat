import pytest

from app.services.kpi import (
    FormulaError,
    detect_tables,
    evaluate_formula,
    validate_arithmetic_consistency,
    validate_formula_shape,
)


def test_basic_arithmetic() -> None:
    assert evaluate_formula("2 + 3 * 4", {}) == 14
    assert evaluate_formula("a + b", {"a": 10, "b": 5}) == 15


def test_functions() -> None:
    assert evaluate_formula("min(a, b)", {"a": 1, "b": 2}) == 1
    assert evaluate_formula("avg(xs)", {"xs": [2, 4, 6]}) == 4
    assert evaluate_formula("round(x, 2)", {"x": 3.14159}) == 3.14


def test_rejects_function_call_on_unknown() -> None:
    with pytest.raises(FormulaError):
        evaluate_formula("exec('import os')", {})


def test_rejects_attribute_access() -> None:
    with pytest.raises(FormulaError):
        evaluate_formula("().__class__", {})


def test_rejects_unknown_variable() -> None:
    with pytest.raises(FormulaError):
        evaluate_formula("mystery", {})


def test_rejects_syntax_error() -> None:
    with pytest.raises(FormulaError):
        evaluate_formula("2 +", {})


def test_validation_shape_flags_non_numeric() -> None:
    issues = validate_formula_shape("a", {"a": "hello"})
    assert any(i.layer == "type" for i in issues)


def test_validation_arithmetic_division_by_zero() -> None:
    issues = validate_arithmetic_consistency("a / b", [{"a": 1, "b": 0}])
    assert any("division by zero" in i.message for i in issues)


def test_validation_arithmetic_non_finite() -> None:
    issues = validate_arithmetic_consistency("x", [{"x": float("inf")}])
    assert any(i.level == "warn" for i in issues)


def test_boolean_ops() -> None:
    assert evaluate_formula("a > 0 and b < 10", {"a": 1, "b": 5}) is True
    assert evaluate_formula("not (x == y)", {"x": 1, "y": 2}) is True


def test_detect_tables_on_two_blocks() -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    # First table
    ws["A1"] = "Q"
    ws["B1"] = "USD"
    ws["A2"], ws["B2"] = "Q1", 10
    ws["A3"], ws["B3"] = "Q2", 12
    # Blank row
    # Second table a few rows down
    ws["A6"] = "Region"
    ws["B6"] = "Count"
    ws["A7"], ws["B7"] = "EMEA", 7
    ws["A8"], ws["B8"] = "APAC", 9

    import io as _io

    buf = _io.BytesIO()
    wb.save(buf)
    tables = detect_tables(buf.getvalue())
    assert len(tables) == 2
    assert tables[0].header == ["Q", "USD"]
    assert len(tables[0].rows) == 2
    assert tables[1].header == ["Region", "Count"]
